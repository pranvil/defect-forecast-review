from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Iterable, List, Optional

from openpyxl.styles import Alignment
from openpyxl.utils.cell import column_index_from_string
from openpyxl.workbook.workbook import Workbook

from app.models import ExportForecastRequest, ForecastTeamRow, MilestoneLabel


@dataclass(frozen=True)
class Cell:
    col: int
    row: int


def _cell(addr: str) -> Cell:
    # Minimal parser for addresses like "C12".
    col_str = "".join([c for c in addr if c.isalpha()])
    row_str = "".join([c for c in addr if c.isdigit()])
    return Cell(col=column_index_from_string(col_str.upper()), row=int(row_str))


def _coerce_date(value: str) -> str:
    # Frontend sends "M/D". Template might show dates as Excel dates.
    return value.strip()


def _max_weeks_from_template() -> int:
    # C .. AB inclusive
    return column_index_from_string("AB") - column_index_from_string("C") + 1


def _week_start_col() -> int:
    return column_index_from_string("C")


def _week_end_col() -> int:
    return column_index_from_string("AB")


def _values_pad(values: List[int], length: int) -> List[int]:
    if len(values) >= length:
        return values[:length]
    return values + [0] * (length - len(values))


def _week_to_index(week_key: str) -> Optional[int]:
    # Accept "W2" / "26W2" / "w2"
    s = week_key.strip().upper()
    if "W" not in s:
        return None
    try:
        n = int(s.split("W", 1)[1])
        if n <= 0:
            return None
        return n - 1
    except Exception:
        return None


def _milestones_to_map(milestones: Iterable[MilestoneLabel]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for m in milestones:
        w = _week_to_index(m.week)
        if w is None:
            continue
        label = m.label.strip()
        if not label:
            continue
        out[w] = label
    return out


def _milestone_rates_to_map(milestones: Iterable[MilestoneLabel], metric: str) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for m in milestones:
        w = _week_to_index(m.week)
        if w is None:
            continue
        value = getattr(m, metric, None)
        if value is None:
            continue
        text = f"{value:g}%"
        out[w] = f"{out[w]} / {text}" if out.get(w) else text
    return out


def _write_row(ws, start: Cell, values: List[object]) -> None:
    for i, v in enumerate(values):
        ws.cell(row=start.row, column=start.col + i, value=v)


def _clear_week_area(ws, max_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row < 1 or merged.min_row > max_row:
            continue
        if merged.max_col < _week_start_col() or merged.min_col > _week_end_col():
            continue
        ws.unmerge_cells(str(merged))
    for row in range(1, max_row + 1):
        for col in range(_week_start_col(), _week_end_col() + 1):
            ws.cell(row=row, column=col).value = None


def _month_label(week_label: str, date_text: str) -> str:
    year = 2026
    clean_week = week_label.strip().upper()
    if "W" in clean_week:
        left = clean_week.split("W", 1)[0]
        if left:
            try:
                raw_year = int(left[-2:])
                year = 2000 + raw_year
            except Exception:
                year = 2026
    month = 0
    clean_date = date_text.strip()
    if "/" in clean_date:
        try:
            month = int(clean_date.split("/", 1)[0])
        except Exception:
            month = 0
    if month <= 0:
        return f"{year}年"
    return f"{year}年{month}月"


def _write_month_header(ws, weekly) -> None:
    if not weekly:
        return
    ws.cell(row=1, column=1, value="项目名")
    ws.cell(row=1, column=2, value="Month")
    segments: list[tuple[str, int, int]] = []
    start_col = _week_start_col()
    for index, point in enumerate(weekly):
        label = _month_label(point.weekLabel, point.date)
        col = start_col + index
        if segments and segments[-1][0] == label:
            prev_label, prev_start, _ = segments[-1]
            segments[-1] = (prev_label, prev_start, col)
        else:
            segments.append((label, col, col))
    for label, first_col, last_col in segments:
        cell = ws.cell(row=1, column=first_col, value=label)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if last_col > first_col:
            ws.merge_cells(start_row=1, start_column=first_col, end_row=1, end_column=last_col)


def _find_row_by_label(ws, label: str, columns: List[int]) -> Optional[int]:
    target = label.strip()
    for r in range(1, ws.max_row + 1):
        for c in columns:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == target:
                return r
    return None


def _ensure_detail_rows(ws, start_row: int, end_row_exclusive: int, desired: int) -> None:
    """
    Ensure there are exactly `desired` rows in [start_row, end_row_exclusive).
    We insert/delete rows so that anchor at end_row_exclusive stays as the first row after the detail block.
    """
    existing = max(0, end_row_exclusive - start_row)
    if existing == desired:
        return
    if existing < desired:
        ws.insert_rows(end_row_exclusive, amount=desired - existing)
    else:
        # delete top-most rows in the block
        ws.delete_rows(start_row, amount=existing - desired)


def fill_forecast_into_template(wb: Workbook, req: ExportForecastRequest) -> None:
    sheet_name = "DRP Plan"
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Template missing sheet: {sheet_name}")
    ws = wb[sheet_name]

    max_weeks = _max_weeks_from_template()
    weekly = req.dataset.weekly[:max_weeks]
    _clear_week_area(ws, ws.max_row)

    # ---- Header (horizontal fixed)
    _write_month_header(ws, weekly)
    ws.cell(row=_cell("A2").row, column=_cell("A2").col, value=req.projectName)
    _write_row(ws, _cell("C2"), [_coerce_date(p.date) for p in weekly])
    _write_row(ws, _cell("C3"), [p.weekLabel for p in weekly])

    # ---- Anchor rows (vertical dynamic)
    total_created_row = _find_row_by_label(ws, "Total Created", columns=[2])
    cum_created_row = _find_row_by_label(ws, "累计创建", columns=[2])
    total_fixed_row = _find_row_by_label(ws, "Total Fixed", columns=[2])
    cum_fixed_row = _find_row_by_label(ws, "累计解决", columns=[2])
    backlog_row = _find_row_by_label(ws, "Backlog", columns=[2])
    mv_row = _find_row_by_label(ws, "MV 版本", columns=[1])

    if not total_created_row or not cum_created_row or not total_fixed_row or not cum_fixed_row or not backlog_row or not mv_row:
        raise KeyError("Template missing one or more anchor labels (Total/Cum/Backlog/MV)")

    created_detail_start = _cell("C3").row + 1  # week row + 1

    # ---- Created detail block (dynamic insert/delete between Week row and Total Created row)
    created_rows: List[ForecastTeamRow] = req.dataset.createdTeams
    _ensure_detail_rows(ws, created_detail_start, total_created_row, desired=len(created_rows))

    # Re-locate anchors after row changes
    total_created_row = _find_row_by_label(ws, "Total Created", columns=[2]) or total_created_row
    cum_created_row = _find_row_by_label(ws, "累计创建", columns=[2]) or cum_created_row
    total_fixed_row = _find_row_by_label(ws, "Total Fixed", columns=[2]) or total_fixed_row
    cum_fixed_row = _find_row_by_label(ws, "累计解决", columns=[2]) or cum_fixed_row
    backlog_row = _find_row_by_label(ws, "Backlog", columns=[2]) or backlog_row
    mv_row = _find_row_by_label(ws, "MV 版本", columns=[1]) or mv_row

    created_values_len = len(weekly)
    for idx, row in enumerate(created_rows):
        r = created_detail_start + idx
        ws.cell(row=r, column=_cell("A4").col, value="Created")
        ws.cell(row=r, column=_cell("B4").col, value=row.team)
        vals = _values_pad(row.values, created_values_len)
        _write_row(ws, Cell(col=_cell("C4").col, row=r), vals)

    # Total / Cum Created (write to anchor rows, horizontal fixed starting at C)
    ws.cell(row=total_created_row, column=_cell("A12").col, value="Created")
    ws.cell(row=total_created_row, column=_cell("B12").col, value="Total Created")
    _write_row(ws, Cell(col=_cell("C12").col, row=total_created_row), [p.created for p in weekly])

    ws.cell(row=cum_created_row, column=_cell("A13").col, value="Created")
    ws.cell(row=cum_created_row, column=_cell("B13").col, value="累计创建")
    _write_row(ws, Cell(col=_cell("C13").col, row=cum_created_row), [p.cumCreated for p in weekly])

    # ---- Fixed detail block (dynamic insert/delete between Cum Created row and Total Fixed row)
    fixed_rows: List[ForecastTeamRow] = req.dataset.fixedTeams
    fixed_detail_start = cum_created_row + 1
    _ensure_detail_rows(ws, fixed_detail_start, total_fixed_row, desired=len(fixed_rows))

    # Re-locate anchors after row changes
    total_fixed_row = _find_row_by_label(ws, "Total Fixed", columns=[2]) or total_fixed_row
    cum_fixed_row = _find_row_by_label(ws, "累计解决", columns=[2]) or cum_fixed_row
    backlog_row = _find_row_by_label(ws, "Backlog", columns=[2]) or backlog_row
    mv_row = _find_row_by_label(ws, "MV 版本", columns=[1]) or mv_row

    for idx, row in enumerate(fixed_rows):
        r = fixed_detail_start + idx
        ws.cell(row=r, column=_cell("A14").col, value="Fixed")
        ws.cell(row=r, column=_cell("B14").col, value=row.team)
        vals = _values_pad(row.values, created_values_len)
        _write_row(ws, Cell(col=_cell("C14").col, row=r), vals)

    # Total / Cum Fixed
    ws.cell(row=total_fixed_row, column=_cell("A23").col, value="Fixed")
    ws.cell(row=total_fixed_row, column=_cell("B23").col, value="Total Fixed")
    _write_row(ws, Cell(col=_cell("C23").col, row=total_fixed_row), [p.fixed for p in weekly])

    ws.cell(row=cum_fixed_row, column=_cell("A24").col, value="Fixed")
    ws.cell(row=cum_fixed_row, column=_cell("B24").col, value="累计解决")
    _write_row(ws, Cell(col=_cell("C24").col, row=cum_fixed_row), [p.cumFixed for p in weekly])

    # ---- Backlog
    ws.cell(row=backlog_row, column=_cell("A25").col, value="遗留")
    ws.cell(row=backlog_row, column=_cell("B25").col, value="Backlog")
    _write_row(ws, Cell(col=_cell("C25").col, row=backlog_row), [p.backlog for p in weekly])

    # ---- Milestones (MV 版本行)
    ws.cell(row=mv_row, column=_cell("A26").col, value="MV 版本")
    milestone_map = _milestones_to_map(req.dataset.milestones)
    mv_values = [""] * len(weekly)
    for i, p in enumerate(weekly):
        w = _week_to_index(p.week)
        if w is None:
            continue
        mv_values[i] = milestone_map.get(w, "")
    _write_row(ws, Cell(col=_cell("C26").col, row=mv_row), mv_values)

    metric_rows = [
        ("问题提交率", "testSubmissionRate"),
        ("问题解决率", "devResolutionRate"),
        ("测试完成率", "testCompletionRate"),
    ]
    for offset, (label, metric) in enumerate(metric_rows, start=1):
        row = mv_row + offset
        if ws.cell(row=row, column=1).value not in (None, "", label):
            ws.insert_rows(row)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value="")
        rate_map = _milestone_rates_to_map(req.dataset.milestones, metric)
        values = []
        for p in weekly:
            w = _week_to_index(p.week)
            values.append(rate_map.get(w, "") if w is not None else "")
        _write_row(ws, Cell(col=_cell("C26").col, row=row), values)

    # Put a tiny export timestamp somewhere safe (optional): keep it minimal.
    # We'll put it in A1 if empty.
    if ws["A1"].value in (None, ""):
        ws["A1"].value = f"Exported at {datetime.now().isoformat(timespec='seconds')}"
