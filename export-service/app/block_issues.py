from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.jira_client import jira_get, jira_post, jira_put, search_issues_paged
from app.models import (
    BlockIssueBatchResult,
    BlockIssueMarkRequest,
    BlockIssueMarkResult,
    BlockIssueRow,
    BlockIssueSearchRequest,
    BlockIssueSearchResult,
    JiraFetchRequest,
)

MAIN_CEA_FIELD = "customfield_11539"
ADDITIONAL_CEA_FIELD = "customfield_11518"
IPR_FIELD = "customfield_13522"
DEADLINE_FIELD = "customfield_11550"
BLOCK_VALUE = "BLOCK"
ALLOWED_BLOCK_STATUSES = {"MORE INFO", "ASSIGNED", "OPENED"}
MAIN_CEA_OPTIONS = {
    "SHOWSTOPPER": {"value": "ShowStopper", "id": "11513"},
    "TOP": {"value": "TOP", "id": "11514"},
    "BLOCK": {"value": "BLOCK", "id": "11515"},
    "BUGASS": {"value": "BUGASS", "id": "11516"},
}
BLOCKER_MAIN_CEA_VALUES = set(MAIN_CEA_OPTIONS)


def _default_deadline() -> str:
    return (date.today() + timedelta(days=7)).isoformat()


def _normalize_key(value: Any) -> str:
    return str(value or "").strip().upper()


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(_string_value(item) for item in value if _string_value(item)).strip()
    if isinstance(value, dict):
        v = value.get("value") or value.get("name") or value.get("id")
        return str(v or "").strip()
    return str(value).strip()


def _date_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.date().isoformat()
    text = str(value).strip()
    return text[:10] if text else ""


def _is_block(value: Any) -> bool:
    if isinstance(value, list):
        return any(_is_block(item) for item in value)
    if isinstance(value, dict):
        return _string_value(value).upper() in BLOCKER_MAIN_CEA_VALUES
    return _string_value(value).upper() in BLOCKER_MAIN_CEA_VALUES


def _resolve_main_cea_option(value: str) -> dict[str, str]:
    normalized = value.strip().upper() or BLOCK_VALUE
    option = MAIN_CEA_OPTIONS.get(normalized)
    if not option:
        allowed = ", ".join(option["value"] for option in MAIN_CEA_OPTIONS.values())
        raise ValueError(f"Main CEA Comment 只能填写：{allowed}")
    return option


def _to_fetch_request(req: BlockIssueMarkRequest | BlockIssueSearchRequest) -> JiraFetchRequest:
    return JiraFetchRequest(
        projectKey=req.projectKey,
        startWeek=req.startWeek,
        endWeek=req.endWeek,
        pullMode=req.pullMode,
        jql=req.jql,
        startDate=req.startDate,
        endDate=req.endDate,
        mode=req.mode,
        baseUrl=req.baseUrl,
        authType=req.authType,
        username=req.username,
        token=req.token,
        verifySsl=req.verifySsl,
        timeoutSec=req.timeoutSec,
    )


def _issue_to_row(issue: dict[str, Any]) -> BlockIssueRow:
    fields = issue.get("fields") if isinstance(issue.get("fields"), dict) else {}
    ipr = fields.get(IPR_FIELD)
    return BlockIssueRow(
        key=str(issue.get("key") or ""),
        summary=_string_value(fields.get("summary")),
        ipr=float(ipr) if isinstance(ipr, (int, float)) else None,
        mainCeaComment=_string_value(fields.get(MAIN_CEA_FIELD)),
        additionalCeaComment=_string_value(fields.get(ADDITIONAL_CEA_FIELD)),
        deadline=_date_value(fields.get(DEADLINE_FIELD)),
    )


def search_block_issues(req: BlockIssueSearchRequest) -> BlockIssueSearchResult:
    project_key = req.projectKey.strip().upper()
    jql = (
        f'project = {project_key} '
        'AND issuetype in (Defect, defect_new) '
        'AND status in ("MORE INFO", "ASSIGNED", "OPENED") '
        "ORDER BY priority ASC, updated DESC"
    )
    fields = ["summary", MAIN_CEA_FIELD, ADDITIONAL_CEA_FIELD, IPR_FIELD, DEADLINE_FIELD]
    issues = search_issues_paged(_to_fetch_request(req), jql, fields=fields)
    rows = [_issue_to_row(issue) for issue in issues if _is_block((issue.get("fields") or {}).get(MAIN_CEA_FIELD))]
    return BlockIssueSearchResult(projectKey=project_key, jql=jql, total=len(rows), issues=rows)


def _read_existing_issue_state(req: BlockIssueMarkRequest) -> tuple[Any, str]:
    fetch_req = _to_fetch_request(req)
    issue_key = req.issueKey.strip().upper()
    data = jira_get(fetch_req, f"rest/api/2/issue/{issue_key}", {"fields": f"{MAIN_CEA_FIELD},status"})
    fields = data.get("fields") if isinstance(data.get("fields"), dict) else {}
    status = fields.get("status") if isinstance(fields.get("status"), dict) else {}
    status_name = _string_value(status.get("name") if isinstance(status, dict) else status)
    return fields.get(MAIN_CEA_FIELD), status_name


def mark_block_issue(req: BlockIssueMarkRequest) -> BlockIssueMarkResult:
    issue_key = req.issueKey.strip().upper()
    if not issue_key:
        return BlockIssueMarkResult(issueKey="", status="failed", message="Issue key 不能为空")
    existing, status_name = _read_existing_issue_state(req)
    if status_name.upper() not in ALLOWED_BLOCK_STATUSES:
        allowed = ", ".join(sorted(ALLOWED_BLOCK_STATUSES))
        return BlockIssueMarkResult(
            issueKey=issue_key,
            status="skipped",
            message=f"当前状态为 {status_name or '未知'}，不在允许状态范围（{allowed}），已跳过",
        )
    if _is_block(existing):
        return BlockIssueMarkResult(issueKey=issue_key, status="skipped", message="已有 Main CEA Comment 标记，已跳过")

    deadline = req.deadline.strip() or _default_deadline()
    main_option = _resolve_main_cea_option(req.mainCeaComment)
    fields: dict[str, Any] = {
        MAIN_CEA_FIELD: {"id": main_option["id"]},
        ADDITIONAL_CEA_FIELD: req.additionalCeaComment.strip(),
        DEADLINE_FIELD: deadline,
    }
    fetch_req = _to_fetch_request(req)
    jira_put(fetch_req, f"rest/api/2/issue/{issue_key}", {"fields": fields})

    comment_status = "not_requested"
    message = f"已添加 {main_option['value']} 标记"
    comment = req.comment.strip()
    if comment:
        try:
            jira_post(fetch_req, f"rest/api/2/issue/{issue_key}/comment", {"body": comment})
            comment_status = "added"
        except ValueError as e:
            comment_status = "failed"
            message = f"字段已更新，但评论添加失败：{e}"
    return BlockIssueMarkResult(issueKey=issue_key, status="updated", message=message, commentStatus=comment_status)


def _batch_req_from_row(
    base: JiraFetchRequest,
    issue_key: str,
    reason: str,
    main_cea: str,
    additional: str,
    deadline: str,
) -> BlockIssueMarkRequest:
    return BlockIssueMarkRequest(
        projectKey=base.projectKey,
        startWeek=base.startWeek,
        endWeek=base.endWeek,
        pullMode=base.pullMode,
        jql=base.jql,
        startDate=base.startDate,
        endDate=base.endDate,
        mode=base.mode,
        baseUrl=base.baseUrl,
        authType=base.authType,
        username=base.username,
        token=base.token,
        verifySsl=base.verifySsl,
        timeoutSec=base.timeoutSec,
        issueKey=issue_key,
        mainCeaComment=main_cea,
        additionalCeaComment=additional,
        deadline=deadline or _default_deadline(),
        comment=reason,
    )


def batch_mark_block_issues(base_req: JiraFetchRequest, file_bytes: bytes) -> BlockIssueBatchResult:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    results: list[BlockIssueMarkResult] = []
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        issue_key = _normalize_key(row[0] if len(row) > 0 else "")
        if not issue_key:
            continue
        total_rows += 1
        reason = _string_value(row[2] if len(row) > 2 else "")
        main_cea = _string_value(row[3] if len(row) > 3 else "") or BLOCK_VALUE
        additional = _string_value(row[4] if len(row) > 4 else "")
        deadline = _date_value(row[5] if len(row) > 5 else "") or _default_deadline()
        try:
            results.append(mark_block_issue(_batch_req_from_row(base_req, issue_key, reason, main_cea, additional, deadline)))
        except Exception as e:
            results.append(BlockIssueMarkResult(issueKey=issue_key, status="failed", message=str(e)))
    return BlockIssueBatchResult(
        totalRows=total_rows,
        updated=sum(1 for x in results if x.status == "updated"),
        skipped=sum(1 for x in results if x.status == "skipped"),
        failed=sum(1 for x in results if x.status == "failed"),
        results=results,
    )


def build_block_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Block Issues"
    headers = ["KEY", "Summary", "Block理由", "Main CEA Comment", "Additional CEA Comment", "Deadline"]
    ws.append(headers)
    ws.append(["MNTNPOM-123", "示例问题摘要", "阻塞原因示例", "BLOCK", "LE1", _default_deadline()])
    header_fill = PatternFill("solid", fgColor="1E293B")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = [18, 42, 46, 22, 24, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F200"
    main_validation = DataValidation(type="list", formula1='"ShowStopper,TOP,BLOCK,BUGASS"', allow_blank=False)
    main_validation.error = "Main CEA Comment 只能选择 ShowStopper、TOP、BLOCK、BUGASS"
    main_validation.errorTitle = "选项错误"
    ws.add_data_validation(main_validation)
    main_validation.add("D2:D200")
    date_validation = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2035,12,31)")
    date_validation.error = "Deadline 请填写日期"
    date_validation.errorTitle = "日期格式错误"
    ws.add_data_validation(date_validation)
    date_validation.add("F2:F200")
    for row in ws.iter_rows(min_row=2, max_row=200, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
